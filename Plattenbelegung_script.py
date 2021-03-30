import openpyxl
import os
from itertools import chain

#Create 384 plate and write Position

Platte384 = openpyxl.Workbook() #Create empty workbook

Sheet = Platte384.get_sheet_by_name("Sheet") #Get Sheet object
Sheet["A1"] = "Position" #Set header for column 1
Sheet["B1"] = "Name" #Set header for column 2

def char_range(c1, c2): #Function to get letters ranging from A to P
    for c in range(ord(c1), ord(c2)+1):
        yield chr(c)

list = [] #Empty list for column 1

for c in char_range("A", "P"): #For loop to create 384 plate position column
    Letter = c
    for i in range(1, 25):
        list.append(Letter + str(i))

#row = 2
for i, value in enumerate(list): #For loop to add plate positions to column 1
    i=i+1    
    Sheet.cell(column=1, row=i+1, value=value)

# Read in excel files and sheet objects

Platte1 = openpyxl.load_workbook("Platte1.xlsx")
Sheet1 = Platte1.get_sheet_by_name("Tabelle1")

Platte2 = openpyxl.load_workbook("Platte2.xlsx")
Sheet2 = Platte2.get_sheet_by_name("Tabelle1")

Platte3 = openpyxl.load_workbook("Platte3.xlsx")
Sheet3 = Platte3.get_sheet_by_name("Tabelle1")

Platte4 = openpyxl.load_workbook("Platte4.xlsx")
Sheet4 = Platte4.get_sheet_by_name("Tabelle1")

def getwriteNames(sheet1, sheet2, sheet3, sheet4):
    name_list_1 = []
    name_list_2 = []
    name_list_3 = []
    name_list_4 = []
    for columnOfCellObjects in sheet1["B2":"B97"]: #Write names of 96 Sheet to empty list
        for cellObj in columnOfCellObjects:
            name_list_1 += [cellObj.value]
    for columnOfCellObjects in sheet2["B2":"B97"]:
        for cellObj in columnOfCellObjects:
            name_list_2 += [cellObj.value]
    for columnOfCellObjects in sheet3["B2":"B97"]: #Write names of 96 Sheet to empty list
        for cellObj in columnOfCellObjects:
            name_list_3 += [cellObj.value]
    for columnOfCellObjects in sheet4["B2":"B97"]:
        for cellObj in columnOfCellObjects:
            name_list_4 += [cellObj.value]
    row_list384_1 = []
    row_list384_2 = []
    row_list384_3 = []
    row_list384_4 = [] 
    for i, value in enumerate(list):
        if (i % 2 == 0 and i != 0 and i in chain(range(2, 26), range(50, 74), range(98, 122), range(146, 170), range(194, 218), range(242, 266), range(290, 314), range(338, 362))):
            row_list384_1.append(i)
        elif (i % 2 != 0 and i != 0 and i in chain(range(2, 26), range(50, 74), range(98, 122), range(146, 170), range(194, 218), range(242, 266), range(290, 314), range(338, 362))):
            row_list384_2.append(i)
    for i, value in enumerate(row_list384_1): #Write names from 96 list to 384 excel sheet
        Sheet.cell(column = 2, row = row_list384_1[i], value = name_list_1[i])
    for i, value in enumerate(row_list384_2): #Write names from 96 list to 384 excel sheet
        Sheet.cell(column = 2, row = row_list384_2[i], value = name_list_2[i])

    for i in range(1, 386):        #Write row numbers from 384 sheet to which names should be pasted into empty
        if (i % 2 == 0 and i != 0 and i in chain(range(26, 50), range(74, 98), range(122, 146), range(170, 194), range(218, 242), range(266, 290), range(314, 338), range(362, 386))):
            row_list384_3.append(i)
        elif (i % 2 != 0 and i != 0 and i in chain(range(26, 50), range(74, 98), range(122, 146), range(170, 194), range(218, 242), range(266, 290), range(314, 338), range(362, 386))):
            row_list384_4.append(i)
    for i, value in enumerate(row_list384_3): #Write names from 96 list to 384 excel sheet
        Sheet.cell(column = 2, row = row_list384_3[i], value = name_list_3[i])
    for i, value in enumerate(row_list384_4): #Write names from 96 list to 384 excel sheet
        Sheet.cell(column = 2, row = row_list384_4[i], value = name_list_4[i])    
   

getwriteNames(Sheet1, Sheet2, Sheet3, Sheet4)

Platte384.save("Platte384.xlsx")    



    
